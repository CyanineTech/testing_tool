import os
import sys
import re
import requests
import argparse
from urllib.parse import urlencode
from openpyxl import Workbook, load_workbook
from configparser import ConfigParser

# 基础配置（仅保留目录相关，配置项从ini读取）
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, 'config.ini')


def find_value(obj, key):
    """递归查找字典/列表中的key值，返回第一个匹配结果"""
    if isinstance(obj, dict):
        if key in obj:
            return obj[key]
        for v in obj.values():
            res = find_value(v, key)
            if res is not None:
                return res
    elif isinstance(obj, list):
        for item in obj:
            res = find_value(item, key)
            if res is not None:
                return res
    return None


def load_config(path):
    """从INI配置文件读取所有必要配置，缺少则中文报错"""
    if not os.path.exists(path):
        print("❌ 错误：未找到配置文件 config.ini！", file=sys.stderr)
        print("📝 请在程序目录下创建 config.ini 文件并配置相关参数", file=sys.stderr)
        sys.exit(1)
    
    config = ConfigParser()
    try:
        config.read(path, encoding='utf-8')
    except Exception as e:
        print(f"❌ 错误：读取配置文件失败 - {str(e)}", file=sys.stderr)
        sys.exit(1)
    
    # 存储配置结果
    config_data = {}
    
    # 1. 读取base section（token必填）
    if not config.has_section('base'):
        print("❌ 错误：config.ini 中缺少 [base] 配置段！", file=sys.stderr)
        print("📝 请在 config.ini 中添加：", file=sys.stderr)
        print("[base]", file=sys.stderr)
        print("token = 你的Token字符串", file=sys.stderr)
        sys.exit(1)
    
    if not config.has_option('base', 'token'):
        print("❌ 错误：[base] 配置段中缺少 token 参数！", file=sys.stderr)
        print("📝 请在 [base] 下添加：token = 你的Token字符串", file=sys.stderr)
        sys.exit(1)
    
    token = config.get('base', 'token').strip()
    if not token:
        print("❌ 错误：token 参数不能为空！", file=sys.stderr)
        print("📝 请在 [base] 下填写有效的 token 字符串", file=sys.stderr)
        sys.exit(1)
    config_data['token'] = token
    
    # 2. 读取service section（host、port必填）
    if not config.has_section('service'):
        print("❌ 错误：config.ini 中缺少 [service] 配置段！", file=sys.stderr)
        print("📝 请在 config.ini 中添加：", file=sys.stderr)
        print("[service]", file=sys.stderr)
        print("host = 你的服务端主机名或IP", file=sys.stderr)
        print("port = 服务端端口号（如9990）", file=sys.stderr)
        sys.exit(1)
    
    # 读取host
    if not config.has_option('service', 'host'):
        print("❌ 错误：[service] 配置段中缺少 host 参数！", file=sys.stderr)
        print("📝 请在 [service] 下添加：host = 你的服务端主机名或IP", file=sys.stderr)
        sys.exit(1)
    host = config.get('service', 'host').strip()
    if not host:
        print("❌ 错误：host 参数不能为空！", file=sys.stderr)
        print("📝 请在 [service] 下填写有效的主机名或IP", file=sys.stderr)
        sys.exit(1)
    config_data['host'] = host
    
    # 读取port
    if not config.has_option('service', 'port'):
        print("❌ 错误：[service] 配置段中缺少 port 参数！", file=sys.stderr)
        print("📝 请在 [service] 下添加：port = 服务端端口号（如9990）", file=sys.stderr)
        sys.exit(1)
    port = config.get('service', 'port').strip()
    if not port:
        print("❌ 错误：port 参数不能为空！", file=sys.stderr)
        print("📝 请在 [service] 下填写有效的端口号（整数）", file=sys.stderr)
        sys.exit(1)
    # 验证端口格式
    if not port.isdigit():
        print("❌ 错误：port 参数必须是整数！", file=sys.stderr)
        print("📝 请在 [service] 下填写有效的端口号（如9990）", file=sys.stderr)
        sys.exit(1)
    config_data['port'] = port
    
    # 3. 读取map section（scene_id必填）
    if not config.has_section('map'):
        print("❌ 错误：config.ini 中缺少 [map] 配置段！", file=sys.stderr)
        print("📝 请在 config.ini 中添加：", file=sys.stderr)
        print("[map]", file=sys.stderr)
        print("scene_id = 场景ID（整数）", file=sys.stderr)
        sys.exit(1)
    
    # 读取scene_id
    if not config.has_option('map', 'scene_id'):
        print("❌ 错误：[map] 配置段中缺少 scene_id 参数！", file=sys.stderr)
        print("📝 请在 [map] 下添加：scene_id = 场景ID（整数）", file=sys.stderr)
        sys.exit(1)
    try:
        scene_id = config.getint('map', 'scene_id')
    except ValueError:
        print("❌ 错误：scene_id 参数必须是整数！", file=sys.stderr)
        print("📝 请在 [map] 下填写有效的整数场景ID", file=sys.stderr)
        sys.exit(1)
    config_data['scene_id'] = scene_id
    
    # 4. 读取excel section（xlsx_path必填）
    if not config.has_section('excel'):
        print("❌ 错误：config.ini 中缺少 [excel] 配置段！", file=sys.stderr)
        print("📝 请在 config.ini 中添加：", file=sys.stderr)
        print("[excel]", file=sys.stderr)
        print("xlsx_path = Excel输出路径（如 ./locations.xlsx）", file=sys.stderr)
        sys.exit(1)
    
    if not config.has_option('excel', 'xlsx_path'):
        print("❌ 错误：[excel] 配置段中缺少 xlsx_path 参数！", file=sys.stderr)
        print("📝 请在 [excel] 下添加：xlsx_path = Excel输出路径", file=sys.stderr)
        sys.exit(1)
    
    xlsx_path = config.get('excel', 'xlsx_path').strip()
    if not xlsx_path:
        print("❌ 错误：xlsx_path 参数不能为空！", file=sys.stderr)
        print("📝 请在 [excel] 下填写有效的Excel输出路径", file=sys.stderr)
        sys.exit(1)
    # 处理相对路径（转为绝对路径）
    if not os.path.isabs(xlsx_path):
        xlsx_path = os.path.join(BASE_DIR, xlsx_path)
    config_data['xlsx_path'] = xlsx_path
    
    return config_data


def fetch_locations(host, port, token, scene_id):
    """从服务端获取Locations数据，含HTTP错误处理（使用配置的port）"""
    url = f"http://{host}:{port}/map_server/locations/"
    params = {'scene_id': scene_id}
    headers = {'Authorization': f'Bearer {token}'}
    try:
        resp = requests.get(url, params=params, headers=headers, timeout=15)
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.ConnectionError:
        print(f"❌ 连接失败：主机 '{host}:{port}' 不可达或服务未启动", file=sys.stderr)
        sys.exit(1)
    except requests.exceptions.Timeout:
        print(f"❌ 请求超时：连接主机 '{host}:{port}' 超过15秒", file=sys.stderr)
        sys.exit(1)
    except requests.exceptions.HTTPError as e:
        status_code = resp.status_code
        if status_code == 401:
            print(f"❌ 认证失败：Token无效或已过期", file=sys.stderr)
        else:
            print(f"❌ HTTP错误 {status_code}：{e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"❌ 获取数据失败：{str(e)}", file=sys.stderr)
        sys.exit(1)


def process_alias(alias):
    """处理别名规则，生成alias_kept"""
    if alias is None:
        return ""
    s = str(alias)
    
    # 保留第一个'-'前 + 第二个'-'后的内容
    first_dash = s.find('-')
    if first_dash == -1:
        kept = s
    else:
        second_dash = s.find('-', first_dash + 1)
        kept = s[:first_dash] + (s[second_dash:] if second_dash != -1 else "")
    
    # 清理冗余内容
    kept = kept.replace('0_copy', '')
    kept = kept.strip('_- ')
    kept = re.sub(r'[_-]{2,}', lambda m: m.group(0)[0], kept)
    return kept


def split_alias_for_sort(alias_kept):
    """
    修复排序逻辑：按「区域数字+序号数字」排序（支持1-24、10-1这种格式）
    排序优先级：
    1. 区域前缀中的数字（如1、2、10，按数字升序）
    2. 序号数字（如24、1，按数字升序）
    3. 纯字符串前缀（如hoist_L，按字母序后置）
    """
    if not alias_kept:
        return (999, 999)  # 空值后置
    
    # 按最后一个'-'拆分（区域前缀 + 序号）
    parts = alias_kept.rsplit('-', 1)
    area_part = parts[0].strip()
    seq_part = parts[1].strip() if len(parts) == 2 else ""
    
    # 提取区域前缀中的数字（核心修复：优先按数字排序）
    area_num = re.findall(r'^\d+', area_part)  # 匹配前缀的纯数字（如"10"从"10-1"中提取）
    if area_num:
        area_sort_key = int(area_num[0])  # 数字区域（如1、2、10）
    else:
        area_sort_key = 999  # 非数字区域（如hoist_L）后置
    
    # 提取序号中的数字
    try:
        seq_sort_key = int(seq_part) if seq_part else 999
    except ValueError:
        seq_sort_key = 999  # 序号非数字后置
    
    # 补充：如果区域前缀完全相同（含非数字），再按原字符串排序（避免歧义）
    return (area_sort_key, area_part, seq_sort_key)


def write_locations_xlsx(path, items, host, scene_id):
    """
    写入Excel（含去重、按区域数字+序号数字排序）
    表格名格式：{host}_{scene_id}（相同参数覆盖，不同参数新增）
    """
    # 生成动态表格名（替换非法字符，确保Excel兼容）
    sheet_name = f"{host}_{scene_id}".replace('/', '_').replace('\\', '_').replace(':', '_')
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]  # Excel表格名最大31个字符
    
    # 收集有效数据
    data_list = []
    for item in items:
        item_id = find_value(item, 'id')
        alias = find_value(item, 'alias')
        if item_id is None:
            print(f"⚠️  跳过无ID的条目：{item}")
            continue
        data_list.append({
            'id': item_id,
            'alias_kept': process_alias(alias)
        })
    
    # 去重（按ID保留最后一条）
    unique_data = {d['id']: d for d in data_list}
    data_list_unique = list(unique_data.values())
    
    # 排序：按「区域数字→区域字符串→序号数字」排序（核心修复）
    data_list_sorted = sorted(
        data_list_unique,
        key=lambda x: split_alias_for_sort(x['alias_kept'])
    )
    
    # 处理Excel文件：存在则打开，不存在则新建
    if os.path.exists(path):
        wb = load_workbook(path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]  # 覆盖已有表格
    else:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])  # 删除默认工作表
    
    # 新建表格并写入数据
    ws = wb.create_sheet(title=sheet_name)
    ws.cell(row=1, column=1, value='id')
    ws.cell(row=1, column=2, value='alias_kept')
    
    for row_idx, data in enumerate(data_list_sorted, start=2):
        ws.cell(row=row_idx, column=1, value=data['id'])
        ws.cell(row=row_idx, column=2, value=data['alias_kept'])
    
    # 保存文件
    wb.save(path)
    print(f"📋 已写入表格：{sheet_name}（{len(data_list_sorted)} 条数据）")
    return len(data_list), len(data_list_sorted)


def parse_arguments():
    """解析命令行参数（仅保留帮助信息，配置从ini读取）"""
    parser = argparse.ArgumentParser(
        description='📌 从地图服务获取Locations数据并生成排序后的Excel文件',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
使用说明：
  1. 所有配置均从 config.ini 文件读取，无需命令行参数
  2. 请确保 config.ini 包含以下配置段：
     
[base]
token = 你的Token字符串（必填）

[service]
host = 服务端主机名或IP（必填）
port = 服务端端口号（如9990，必填）

[map]
scene_id = 场景ID（整数，必填）

[excel]
xlsx_path = Excel输出路径（如 ./locations.xlsx，必填）

  3. 表格命名规则：
     自动生成格式为「主机名_SceneID」（如 ubuntu-170_18）
     相同参数覆盖表格，不同参数新增表格（不删除原有表格）
        '''
    )
    return parser.parse_args()


def main():
    """主逻辑：参数解析 → 配置加载 → 数据获取 → 处理写入"""
    # 解析命令行参数（仅处理--help）
    args = parse_arguments()
    
    # 加载配置文件（缺少配置会中文报错并退出）
    print("🔍 正在读取配置文件...")
    config = load_config(CONFIG_PATH)
    host = config['host']
    port = config['port']
    scene_id = config['scene_id']
    token = config['token']
    output_path = config['xlsx_path']
    
    # 生成表格名（去掉端口号）
    sheet_name = f"{host}_{scene_id}".replace('/', '_').replace('\\', '_').replace(':', '_')
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    
    print("=" * 60)
    print("📋 执行配置：")
    print(f"   服务地址  : {host}:{port}")
    print(f"   Scene ID  : {scene_id}")
    print(f"   输出路径  : {output_path}")
    print(f"   目标表格  : {sheet_name}（相同参数覆盖，不同参数新增）")
    print(f"   Token     : {'***' + token[-4:] if len(token) >= 8 else token}（已隐藏部分）")
    print(f"   配置文件  : {CONFIG_PATH}")
    print("=" * 60)
    
    # 数据处理流程
    try:
        print("🔄 正在获取Locations数据...")
        data = fetch_locations(host, port, token, scene_id)
        
        # 解析数据结构
        if isinstance(data, dict) and 'results' in data and isinstance(data['results'], list):
            items = data['results']
        elif isinstance(data, list):
            items = data
        else:
            items = next((v for v in data.values() if isinstance(v, list)), None)
            if not items:
                print("❌ 错误：服务端返回数据中未找到有效Locations列表", file=sys.stderr)
                sys.exit(1)
        
        print(f"📊 已获取 {len(items)} 条原始数据，正在处理（去重+数字排序）...")
        # 传入host和scene_id用于生成表格名（去掉port）
        valid_count, final_count = write_locations_xlsx(output_path, items, host, scene_id)
        
        # 输出结果统计
        print("=" * 60)
        print("✅ 任务执行完成！")
        print("📈 数据统计：")
        print(f"   - 原始数据总数    : {len(items)} 条")
        print(f"   - 有效数据（含ID）: {valid_count} 条")
        print(f"   - 去重排序后数据  : {final_count} 条")
        print(f"📁 输出文件：{output_path}")
        
        # 检查是新增还是覆盖
        is_cover = False
        if os.path.exists(output_path):
            wb = load_workbook(output_path)
            is_cover = sheet_name in wb.sheetnames
            wb.close()
        
        print(f"📋 操作结果：{'覆盖' if is_cover else '新增'}表格 {sheet_name}")
        print("=" * 60)
    except Exception as e:
        print(f"❌ 任务执行失败：{str(e)}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()