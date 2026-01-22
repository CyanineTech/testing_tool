import os
import sys
import argparse
import random
import time
import requests
import logging
from datetime import datetime, timedelta
from openpyxl import load_workbook
from typing import Dict, List, Optional, Set, Tuple, Union, Any
from configparser import ConfigParser
import signal
from dataclasses import dataclass, field
from functools import wraps
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import json

# ==============================
# 常量定义
# ==============================
VALID_RULES = {1, 2}
REQUIRED_EXCEL_COLUMNS = {'id', 'alias_kept'}
REQUIRED_CONFIG_SECTIONS = ['base', 'service', 'map', 'business', 'task', 'excel', 'log']
REQUIRED_CONFIG_KEYS = {
    'base': ['account', 'password', 'token'],
    'service': ['host', 'port'],
    'map': ['scene_id'],
    'business': ['rule', 'areas'],
    'task': ['locations'],
    'excel': ['xlsx_path'],
    'log': ['debug', 'log_file']
}
TARGET_ERROR_ID = 50421021  # 目标error_id（仅该ID视为成功）
MAX_CONTINUOUS_FAILURES = 5  # 最大连续失败次数
DEFAULT_ERROR_INFO = "无详细错误信息"  # 默认错误信息

# ==============================
# 数据类定义
# ==============================
@dataclass
class TaskConfig:
    """任务配置数据类（完全从config.ini读取）"""
    # [base] 节
    account: str
    password: str
    token: str
    
    # [service] 节
    host: str
    port: int
    
    # [map] 节
    scene_id: int
    
    # [business] 节
    rule: int
    areas: List[str]
    fixed_store: str
    
    # [task] 节
    locations: List[str]
    
    # [excel] 节
    xlsx_path: str
    sheet_name: Optional[str]
    
    # [log] 节
    debug: bool
    log_file: str
    
    # [request] 节（默认配置生效）
    request_timeout: float = 15.0
    retry_count: int = 0
    retry_delay: float = 1.0

@dataclass
class LocationInfo:
    """位置信息数据类"""
    location_id: str
    area: str
    number: int
    full_area: str

@dataclass
class TaskStats:
    """任务统计数据类"""
    success: int = 0
    fail: int = 0
    total: int = 0
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None
    valid_areas: List[str] = field(default_factory=list)
    total_task_count: int = 0
    skipped: int = 0  # 跳过的任务数
    continuous_failures: int = 0  # 连续失败次数

# ==============================
# 全局变量
# ==============================
global_config: Optional[TaskConfig] = None
task_stats = TaskStats()
resource_handles: Dict[str, Optional[object]] = {"excel_workbook": None, "http_session": None}
is_running = False
console_logger: Optional[logging.Logger] = None
file_logger: Optional[logging.Logger] = None

# ==============================
# 装饰器
# ==============================
def exception_handler(return_value: Union[bool, List, Tuple, None] = False):
    """异常处理装饰器"""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except Exception as e:
                func_name = func.__name__
                # 检查日志是否已初始化
                if file_logger:
                    file_logger.error(f"函数 {func_name} 执行异常", exc_info=True)
                if console_logger:
                    console_logger.error(f"❌ {func_name} 执行失败：{str(e)}")
                else:
                    print(f"❌ {func_name} 执行失败：{str(e)}")
                return return_value
        return wrapper
    return decorator

# ==============================
# 日志相关函数
# ==============================
def init_loggers(log_file: str, debug: bool):
    """初始化双日志：控制台（简化）+ 文件（详细）"""
    global console_logger, file_logger

    # 确保日志目录存在
    log_dir = os.path.dirname(log_file)
    if log_dir and not os.path.exists(log_dir):
        os.makedirs(log_dir)

    # 1. 文件日志（详细，保留所有信息）
    file_handler = logging.FileHandler(log_file, encoding="utf-8", mode="a")
    file_formatter = logging.Formatter(
        "%(asctime)s - %(name)s - %(levelname)s - %(filename)s:%(lineno)d - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    file_handler.setFormatter(file_formatter)
    file_logger = logging.getLogger("TaskFileLogger")
    file_logger.addHandler(file_handler)
    file_logger.setLevel(logging.DEBUG if debug else logging.INFO)
    file_logger.propagate = False

    # 2. 控制台日志（简化，仅关键信息）
    console_handler = logging.StreamHandler(sys.stdout)
    console_formatter = logging.Formatter("%(message)s")
    console_handler.setFormatter(console_formatter)
    console_logger = logging.getLogger("TaskConsoleLogger")
    console_logger.addHandler(console_handler)
    console_logger.setLevel(logging.INFO)
    console_logger.propagate = False

    if file_logger:
        file_logger.info("日志系统初始化完成")

def load_log_config_from_ini(config_path: str) -> Tuple[Optional[str], Optional[bool]]:
    """仅加载日志相关配置（用于初始化日志）"""
    if not os.path.exists(config_path):
        print(f"❌ 配置文件 {config_path} 不存在！")
        return None, None

    try:
        config = ConfigParser()
        config.optionxform = str
        config.read(config_path, encoding="utf-8")

        # 检查log节是否存在
        if not config.has_section("log"):
            print(f"❌ 配置文件缺少 [log] 节！")
            return None, None

        # 读取日志配置
        log_file = config.get("log", "log_file", fallback=None)
        debug_str = config.get("log", "debug", fallback=None)

        if not log_file:
            print(f"❌ [log] 节缺少 log_file 配置！")
            return None, None

        debug = parse_ini_bool(debug_str) if debug_str is not None else False

        return log_file.strip(), debug

    except Exception as e:
        print(f"❌ 读取日志配置失败：{str(e)}")
        return None, None

# ==============================
# 配置解析工具函数
# ==============================
def parse_ini_list(value: str) -> List[str]:
    """解析INI列表配置（去重、过滤空值）"""
    if not value:
        return []
    return list(set([item.strip() for item in value.split(',') if item.strip()]))

def parse_ini_bool(value: str) -> bool:
    """解析INI布尔配置"""
    if not value:
        return False
    return value.strip().lower() in ("yes", "true", "1")

def parse_ini_number(value: str, is_int: bool = True) -> Union[int, float, None]:
    """解析INI数字配置（无默认值，解析失败返回None）"""
    if not value:
        return None
    try:
        return int(value.strip()) if is_int else float(value.strip())
    except (ValueError, TypeError):
        return None

# ==============================
# 配置加载和验证函数
# ==============================
def load_full_config(config_path: str, cli_args: argparse.Namespace) -> Optional[TaskConfig]:
    """加载完整配置（完全从config.ini读取，[request]节默认生效）"""
    print(f"📋 配置文件路径：{config_path}")

    # 检查配置文件是否存在
    if not os.path.exists(config_path):
        print(f"❌ 配置文件 {config_path} 不存在！")
        return None

    try:
        config = ConfigParser()
        config.optionxform = str
        config.read(config_path, encoding="utf-8")

        # 1. 验证必填节是否存在
        missing_sections = [sec for sec in REQUIRED_CONFIG_SECTIONS if not config.has_section(sec)]
        if missing_sections:
            print(f"❌ 配置文件缺少必填节：{missing_sections}")
            return None

        # 2. 读取并验证各节配置
        config_dict = {}

        # [base] 节
        base_config = {}
        for key in REQUIRED_CONFIG_KEYS['base']:
            value = config.get("base", key, fallback="").strip()
            if not value:
                print(f"❌ [base] 节缺少必填配置：{key}")
                return None
            base_config[key] = value
        config_dict.update(base_config)

        # [service] 节（host必填，读取后输出）
        service_config = {}
        for key in REQUIRED_CONFIG_KEYS['service']:
            value = config.get("service", key, fallback="").strip()
            if not value:
                print(f"❌ [service] 节缺少必填配置：{key}")
                return None
            if key == 'port':
                port = parse_ini_number(value, is_int=True)
                if port is None or not (1 <= port <= 65535):
                    print(f"❌ [service] 节 port 配置无效：{value}（需1-65535之间的整数）")
                    return None
                service_config[key] = port
            else:
                service_config[key] = value
        config_dict.update(service_config)

        # 输出当前使用的主机地址
        print(f"🔌 当前接口调用主机地址：{service_config['host']}（端口：{service_config['port']}）")
        print(f"✅ 主机配置加载完成")

        # [map] 节
        map_config = {}
        for key in REQUIRED_CONFIG_KEYS['map']:
            value = config.get("map", key, fallback="").strip()
            if not value:
                print(f"❌ [map] 节缺少必填配置：{key}")
                return None
            scene_id = parse_ini_number(value, is_int=True)
            if scene_id is None or scene_id <= 0:
                print(f"❌ [map] 节 scene_id 配置无效：{value}（需正整数）")
                return None
            map_config[key] = scene_id
        config_dict.update(map_config)

        # [business] 节
        business_config = {}
        for key in REQUIRED_CONFIG_KEYS['business']:
            value = config.get("business", key, fallback="").strip()
            if not value:
                print(f"❌ [business] 节缺少必填配置：{key}")
                return None
            if key == 'rule':
                rule = parse_ini_number(value, is_int=True)
                if rule not in VALID_RULES:
                    print(f"❌ [business] 节 rule 配置无效：{value}（仅支持{VALID_RULES}）")
                    return None
                business_config[key] = rule
            elif key == 'areas':
                areas = parse_ini_list(value)
                if not areas:
                    print(f"❌ [business] 节 areas 配置无效（不能为空列表）")
                    return None
                business_config[key] = areas
        # 可选配置：fixed_store
        fixed_store = config.get("business", "fixed_store", fallback="").strip()
        business_config['fixed_store'] = fixed_store
        config_dict.update(business_config)

        # [task] 节
        task_config = {}
        for key in REQUIRED_CONFIG_KEYS['task']:
            value = config.get("task", key, fallback="").strip()
            if not value:
                print(f"❌ [task] 节缺少必填配置：{key}")
                return None
            locations = parse_ini_list(value)
            if not locations:
                print(f"❌ [task] 节 locations 配置无效（不能为空列表）")
                return None
            task_config[key] = locations
        config_dict.update(task_config)

        # [excel] 节
        excel_config = {}
        for key in REQUIRED_CONFIG_KEYS['excel']:
            value = config.get("excel", key, fallback="").strip()
            if not value:
                print(f"❌ [excel] 节缺少必填配置：{key}")
                return None
            excel_config[key] = value
        # 可选配置：sheet_name
        sheet_name = config.get("excel", "sheet_name", fallback="").strip()
        excel_config['sheet_name'] = sheet_name if sheet_name else None
        config_dict.update(excel_config)

        # [log] 节（已在日志初始化时验证过）
        log_config = {}
        log_file = config.get("log", "log_file", fallback="").strip()
        debug_str = config.get("log", "debug", fallback="false").strip()
        log_config['log_file'] = log_file
        log_config['debug'] = parse_ini_bool(debug_str)
        config_dict.update(log_config)

        # [request] 节（默认生效，优先读取配置文件，无配置则用默认值）
        request_config = {
            'request_timeout': 15.0,
            'retry_count': 0,
            'retry_delay': 1.0
        }
        if config.has_section("request"):
            timeout_str = config.get("request", "timeout", fallback="").strip()
            if timeout_str:
                timeout = parse_ini_number(timeout_str, is_int=False)
                if timeout is not None and timeout > 0:
                    request_config['request_timeout'] = timeout

            retry_count_str = config.get("request", "retry_count", fallback="").strip()
            if retry_count_str:
                retry_count = parse_ini_number(retry_count_str, is_int=True)
                if retry_count is not None and retry_count >= 0:
                    request_config['retry_count'] = retry_count

            retry_delay_str = config.get("request", "retry_delay", fallback="").strip()
            if retry_delay_str:
                retry_delay = parse_ini_number(retry_delay_str, is_int=False)
                if retry_delay is not None and retry_delay >= 0:
                    request_config['retry_delay'] = retry_delay
        config_dict.update(request_config)

        # 3. 应用命令行参数覆盖（如果有）
        if cli_args.rule is not None:
            if cli_args.rule in VALID_RULES:
                print(f"⚠️  命令行参数覆盖 [business] 节 rule：{cli_args.rule}")
                config_dict['rule'] = cli_args.rule
            else:
                print(f"❌ 命令行参数 rule 无效（仅支持{VALID_RULES}）")
                return None

        if cli_args.areas is not None:
            print(f"⚠️  命令行参数覆盖 [business] 节 areas：{cli_args.areas}")
            config_dict['areas'] = cli_args.areas

        if cli_args.fixed_store is not None:
            print(f"⚠️  命令行参数覆盖 [business] 节 fixed_store：{cli_args.fixed_store}")
            config_dict['fixed_store'] = cli_args.fixed_store

        if cli_args.debug:
            print(f"⚠️  命令行参数开启 debug 模式")
            config_dict['debug'] = True

        # 4. 最终验证配置逻辑
        if config_dict['rule'] == 1 and len(config_dict['areas']) != 1:
            print(f"❌ 规则1要求 areas 仅含1个区域，当前：{config_dict['areas']}")
            return None

        if config_dict['rule'] == 2 and len(config_dict['areas']) < 2:
            print(f"❌ 规则2要求 areas 至少含2个区域，当前：{config_dict['areas']}")
            return None

        if config_dict['fixed_store'] and config_dict['fixed_store'] not in config_dict['locations']:
            print(f"❌ 固定store {config_dict['fixed_store']} 不在 locations 列表中")
            return None

        # 5. 转换为配置对象
        task_config = TaskConfig(
            # [base]
            account=config_dict['account'],
            password=config_dict['password'],
            token=config_dict['token'],
            # [service]
            host=config_dict['host'],
            port=config_dict['port'],
            # [map]
            scene_id=config_dict['scene_id'],
            # [business]
            rule=config_dict['rule'],
            areas=config_dict['areas'],
            fixed_store=config_dict['fixed_store'],
            # [task]
            locations=config_dict['locations'],
            # [excel]
            xlsx_path=config_dict['xlsx_path'],
            sheet_name=config_dict['sheet_name'],
            # [log]
            debug=config_dict['debug'],
            log_file=config_dict['log_file'],
            # [request]
            request_timeout=config_dict['request_timeout'],
            retry_count=config_dict['retry_count'],
            retry_delay=config_dict['retry_delay']
        )

        # 日志已初始化时输出配置信息
        if file_logger:
            file_logger.info(f"配置加载成功：{task_config}")
        else:
            print(f"✅ 所有配置加载完成")

        return task_config

    except Exception as e:
        print(f"❌ 加载配置失败：{str(e)}")
        return None

# ==============================
# Excel相关函数
# ==============================
@exception_handler(return_value=[])
def load_xlsx_data(xlsx_path: str, sheet_name: Optional[str]) -> List[LocationInfo]:
    """加载Excel数据（未指定工作表时使用最新工作表（最后一个））"""
    console_logger.info(f"📊 加载Excel数据：{xlsx_path}")
    file_logger.info(f"开始加载Excel文件：{xlsx_path}，指定工作表：{sheet_name or '无（将使用最新工作表）'}")

    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Excel文件不存在：{xlsx_path}")

    # 验证文件格式
    if not xlsx_path.endswith(('.xlsx', '.xlsm')):
        raise ValueError(f"不支持的文件格式：{xlsx_path}（仅支持.xlsx/.xlsm）")

    try:
        # 只读模式打开，提升性能
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        resource_handles["excel_workbook"] = wb

        # 选择工作表：未指定时使用最后一个（最新）工作表
        if not sheet_name:
            if len(wb.sheetnames) == 0:
                raise ValueError("Excel文件中无可用工作表")
            sheet_name = wb.sheetnames[-1]  # 取最后一个工作表（最新）
            console_logger.info(f"📑 未指定工作表，使用最新工作表（最后一个）：{sheet_name}")
            file_logger.info(f"未指定工作表，自动选择最新工作表（最后一个）：{sheet_name}，所有工作表：{wb.sheetnames}")
        
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在，可用工作表：{wb.sheetnames}")
        
        ws = wb[sheet_name]
        file_logger.info(f"成功打开工作表：{sheet_name}（总行数：{ws.max_row}）")

        # 解析表头
        headers = []
        for cell in ws[1]:
            header_value = cell.value.strip() if cell.value and isinstance(cell.value, str) else str(cell.value) if cell.value else None
            headers.append(header_value)
        
        # 验证必要列
        missing_cols = REQUIRED_EXCEL_COLUMNS - set(headers)
        if missing_cols:
            raise ValueError(f"缺少必要列：{missing_cols}（需包含 {REQUIRED_EXCEL_COLUMNS}）")
        
        id_col_idx = headers.index('id')
        alias_col_idx = headers.index('alias_kept')
        file_logger.info(f"表头解析成功：id列索引={id_col_idx}，alias_kept列索引={alias_col_idx}")

        # 解析数据（分批读取，提升大文件性能）
        location_list = []
        batch_size = 1000
        batch_count = 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过空行
            if all(cell is None for cell in row):
                continue
            
            # 解析location_id
            location_id = row[id_col_idx]
            if location_id is None:
                file_logger.warning(f"第 {row_num} 行：id为空，跳过")
                task_stats.skipped += 1
                continue
            
            location_id = str(location_id).strip()
            if not location_id:
                file_logger.warning(f"第 {row_num} 行：id为空字符串，跳过")
                task_stats.skipped += 1
                continue

            # 解析alias_kept
            alias_kept = row[alias_col_idx] or ""
            alias_kept = str(alias_kept).strip().replace(" - ", "-")
            if not alias_kept:
                file_logger.warning(f"第 {row_num} 行：alias_kept为空，跳过（id：{location_id}）")
                task_stats.skipped += 1
                continue

            # 解析区域和编号（支持字母数字混合区域，如hoist_L、hoist_R）
            parts = alias_kept.rsplit('-', 1)
            if len(parts) == 2 and parts[1].strip().isdigit():
                area = parts[0].strip()
                number = int(parts[1].strip())
            else:
                # 处理无编号的区域（如hoist_L、hoist_R）
                area = alias_kept.strip()
                number = 0  # 用0表示无编号
            
            if not area:
                file_logger.warning(f"第 {row_num} 行：区域为空（{alias_kept}），跳过（id：{location_id}）")
                task_stats.skipped += 1
                continue

            # 添加到列表
            full_area = alias_kept.strip()  # 直接使用原始alias_kept作为完整区域标识
            location_info = LocationInfo(
                location_id=location_id,
                area=area,
                number=number,
                full_area=full_area
            )
            location_list.append(location_info)

            # 批量日志
            batch_count += 1
            if batch_count % batch_size == 0:
                file_logger.info(f"已解析 {len(location_list)} 条数据（当前行：{row_num}）")

        # 统计信息
        parsed_areas = list({item.area for item in location_list})
        console_logger.info(f"✅ Excel加载完成：")
        console_logger.info(f"  - 有效数据：{len(location_list)} 条")
        console_logger.info(f"  - 总行数：{ws.max_row - 1} 行")
        console_logger.info(f"  - 跳过行数：{task_stats.skipped} 行")
        console_logger.info(f"  - 解析区域：{len(parsed_areas)} 个（{sorted(parsed_areas)}）")
        
        file_logger.info(f"Excel加载完成：有效数据{len(location_list)}条，总行数{ws.max_row - 1}行，跳过{task_stats.skipped}行，解析区域{len(parsed_areas)}个")
        return location_list

    except Exception as e:
        if file_logger:
            file_logger.error(f"Excel加载失败：{str(e)}", exc_info=True)
        raise

# ==============================
# 区域处理相关函数
# ==============================
@exception_handler(return_value=(None, None))
def group_and_validate_areas(
    locations: List[LocationInfo], 
    selected_areas: List[str]
) -> Tuple[Optional[Dict[str, List[LocationInfo]]], Optional[List[str]]]:
    """区域分组与验证（支持字母数字混合区域）"""
    console_logger.info(f"\n🔍 区域分组与验证...")
    file_logger.info(f"开始区域验证：配置区域={sorted(selected_areas)}（仅来自[business]节）")

    # 按区域分组（不区分大小写，支持精确匹配）
    grouped = {}
    for loc in locations:
        area_key = loc.area
        if area_key not in grouped:
            grouped[area_key] = []
        grouped[area_key].append(loc)
    
    # 排序并过滤空组
    grouped = {
        area: sorted(locs, key=lambda x: (x.number, x.full_area)) 
        for area, locs in grouped.items() 
        if area and locs
    }

    # 验证逻辑
    available_areas = set(grouped.keys())
    selected_areas_set = set(selected_areas)
    matched_areas = selected_areas_set & available_areas
    unmatched_areas = selected_areas_set - available_areas
    valid_areas = sorted([area for area in matched_areas if len(grouped[area]) > 0])
    empty_areas = sorted([area for area in matched_areas if len(grouped[area]) == 0])

    # 详细日志
    file_logger.info(f"区域验证详情：")
    file_logger.info(f"  - 配置区域（来自[business]节）：{sorted(selected_areas)}")
    file_logger.info(f"  - 可用区域：{sorted(available_areas)}")
    file_logger.info(f"  - 匹配区域：{sorted(matched_areas)}")
    file_logger.info(f"  - 不匹配区域：{sorted(unmatched_areas) if unmatched_areas else '无'}")
    file_logger.info(f"  - 空区域（无任务）：{empty_areas if empty_areas else '无'}")
    file_logger.info(f"  - 有效区域：{valid_areas}")

    # 错误处理
    if unmatched_areas:
        raise ValueError(f"以下区域在Excel中不存在：{sorted(unmatched_areas)}")
    if empty_areas:
        raise ValueError(f"以下区域无任务数据：{empty_areas}")
    if not valid_areas:
        raise ValueError("无有效区域（匹配区域均无任务）")

    # 统计任务数
    total_task_count = sum(len(locs) for locs in grouped.values() if locs[0].area in valid_areas)
    task_stats.total_task_count = total_task_count
    task_stats.valid_areas = valid_areas

    # 控制台输出
    console_logger.info(f"✅ 区域验证通过！")
    console_logger.info(f"  - 有效区域：{len(valid_areas)} 个（{valid_areas}）")
    console_logger.info(f"  - 待执行任务：{total_task_count} 个")
    file_logger.info(f"区域验证通过：有效区域{len(valid_areas)}个，待执行任务{total_task_count}个")

    return grouped, valid_areas

# ==============================
# HTTP请求相关函数
# ==============================
def init_http_session(config: TaskConfig) -> requests.Session:
    """初始化HTTP会话（使用[request]节配置）"""
    session = requests.Session()
    
    # 配置重试策略（使用[request]节的retry_count和retry_delay）
    retry_strategy = Retry(
        total=config.retry_count,
        backoff_factor=config.retry_delay,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["PUT"]
    )
    
    adapter = HTTPAdapter(
        max_retries=retry_strategy,
        pool_connections=10,
        pool_maxsize=10
    )
    
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    
    resource_handles["http_session"] = session
    file_logger.info(f"HTTP会话初始化完成：超时={config.request_timeout}s，重试={config.retry_count}次，延迟={config.retry_delay}s")
    return session

def check_continuous_failures() -> bool:
    """检查连续失败次数是否达到阈值，达到则返回True（需要停止程序）"""
    if task_stats.continuous_failures >= MAX_CONTINUOUS_FAILURES:
        error_msg = f"⚠️  连续{MAX_CONTINUOUS_FAILURES}次接口调用失败，程序已停止！请检查库位ID有效性、接口配置或服务状态。"
        console_logger.error(error_msg)
        file_logger.error(error_msg)
        # 输出最终报告并清理资源
        task_stats.end_time = datetime.now()
        print_final_report()
        cleanup_resources()
        sys.exit(3)  # 连续失败退出码
    return False

@exception_handler(return_value=False)
def send_task(
    session: requests.Session,
    location_id: str,
    store_location_id: str,
    config: TaskConfig
) -> bool:
    """发送接口请求（支持返回校验：error_id == 50421021 才视为成功）"""
    url = f"http://{config.host}:{config.port}/dispatch_server/dispatch/start/location_call/task/"
    headers = {
        "Authorization": f"Bearer {config.token}",
        "Content-Type": "application/json"
    }
    payload = {
        "location_id": location_id,
        "store_location_id": store_location_id,
        "scene_id": config.scene_id
    }

    # 调试日志
    if config.debug and file_logger:
        file_logger.debug(f"发送请求：")
        file_logger.debug(f"  URL：{url}")
        file_logger.debug(f"  Headers：{headers}")
        file_logger.debug(f"  Payload：{payload}")
        file_logger.debug(f"  超时：{config.request_timeout}s，重试：{config.retry_count}次")

    try:
        response = session.put(
            url,
            json=payload,
            headers=headers,
            timeout=config.request_timeout
        )
        response.raise_for_status()

        # 解析响应JSON
        try:
            resp_data = response.json()
        except json.JSONDecodeError as e:
            error_msg = f"响应格式错误（非JSON）：{response.text[:500]}"
            file_logger.error(f"请求失败：{error_msg} | location_id={location_id}，store={store_location_id}")
            # 更新连续失败计数
            task_stats.continuous_failures += 1
            console_logger.error(f"  ❌ 执行失败：{error_msg}")
            console_logger.warning(f"  ⚠️  连续失败次数：{task_stats.continuous_failures}/{MAX_CONTINUOUS_FAILURES}")
            return False

        # 调试日志：输出完整响应
        if config.debug and file_logger:
            file_logger.debug(f"响应数据：{json.dumps(resp_data, ensure_ascii=False, indent=2)}")

        # 校验返回结果
        success = resp_data.get("success", False)
        error_id = resp_data.get("msg", {}).get("detail", {}).get("error_id")
        
        # 处理错误信息（为空时显示默认值）
        error_info = resp_data.get("msg", {}).get("detail", {}).get("info", "")
        # 去除前后空格，判断是否为空
        if not error_info or str(error_info).strip() == "":
            error_info = DEFAULT_ERROR_INFO
        else:
            error_info = str(error_info).strip()

        # 校验逻辑：仅当success为True 或 (success为False但error_id == TARGET_ERROR_ID) 时视为成功
        if success or (error_id is not None and error_id == TARGET_ERROR_ID):
            # 成功日志优化：不显示None的error_id
            success_log_parts = [
                f"请求成功：location_id={location_id}，store={store_location_id}",
                f"状态码={response.status_code}",
                f"success={success}"
            ]
            # 只有error_id存在时才添加到日志
            if error_id is not None:
                success_log_parts.append(f"error_id={error_id}")
            file_logger.info(" | ".join(success_log_parts))
            
            # 重置连续失败计数
            task_stats.continuous_failures = 0
            return True
        else:
            # 失败处理：确保info有默认值，error_id显示为"无"如果是None
            display_error_id = error_id if error_id is not None else "无"
            error_detail = f"{error_info}（error_id：{display_error_id}）"
            
            # 失败日志优化：明确显示空值情况
            file_logger.error(
                f"请求失败：location_id={location_id}，store={store_location_id} | "
                f"状态码={response.status_code}，success={success}，"
                f"error_id={display_error_id}，info={error_info}"
            )
            
            # 更新连续失败计数
            task_stats.continuous_failures += 1
            console_logger.error(f"  ❌ 执行失败：{error_detail}")
            console_logger.warning(f"  ⚠️  连续失败次数：{task_stats.continuous_failures}/{MAX_CONTINUOUS_FAILURES}")
            return False

    except requests.exceptions.RequestException as e:
        error_details = []
        error_details.append(f"错误类型：{type(e).__name__}")
        error_details.append(f"错误信息：{str(e)}")
        
        if hasattr(e, 'response') and e.response is not None:
            error_details.append(f"状态码：{e.response.status_code}")
            error_details.append(f"响应内容：{e.response.text[:500]}")  # 限制长度
        
        error_msg = " | ".join(error_details)
        file_logger.error(
            f"请求失败：location_id={location_id}，store={store_location_id} | {error_msg}",
            exc_info=config.debug  # 调试模式下输出完整堆栈
        )
        # 更新连续失败计数
        task_stats.continuous_failures += 1
        console_logger.error(f"  ❌ 执行失败：网络请求异常 - {str(e)}")
        console_logger.warning(f"  ⚠️  连续失败次数：{task_stats.continuous_failures}/{MAX_CONTINUOUS_FAILURES}")
        return False

# ==============================
# 任务执行相关函数
# ==============================
def run_rule1(
    grouped_areas: Dict[str, List[LocationInfo]],
    valid_areas: List[str],
    config: TaskConfig
):
    """规则1：单个区域顺序执行（包含连续失败检查）"""
    global is_running
    is_running = True
    target_area = valid_areas[0]
    area_locations = grouped_areas[target_area]
    session = init_http_session(config)

    console_logger.info(f"\n🚀 开始执行任务（规则1：区域{target_area}顺序调用）")
    file_logger.info(
        f"开始执行规则1：目标区域={target_area}（来自[business]节areas），任务数={len(area_locations)}，"
        f"store模式={'固定：' + config.fixed_store if config.fixed_store else '随机'}，"
        f"请求配置：超时={config.request_timeout}s，重试={config.retry_count}次，"
        f"成功条件：success=true 或 error_id={TARGET_ERROR_ID}"
    )

    try:
        for idx, loc in enumerate(area_locations, start=1):
            if not is_running:
                break

            # 检查连续失败次数，达到阈值则停止
            if check_continuous_failures():
                return

            # 选择store
            selected_store = config.fixed_store if config.fixed_store else random.choice(config.locations)
            
            # 控制台进度显示
            console_logger.info(f"\n[{idx}/{len(area_locations)}] 区域：{loc.full_area} | ID：{loc.location_id}")
            
            # 发送任务
            success = send_task(session, loc.location_id, selected_store, config)

            # 更新统计
            if success:
                task_stats.success += 1
                console_logger.info(f"  ✅ 执行成功 | 目标store：{selected_store}")
            else:
                task_stats.fail += 1
            task_stats.total += 1

            # 任务间隔
            if idx < len(area_locations):
                time.sleep(0.5)

    except Exception as e:
        console_logger.error(f"\n❌ 任务执行异常：{str(e)}")
        if file_logger:
            file_logger.error(f"规则1执行异常", exc_info=True)
    finally:
        is_running = False

    # 执行结果
    if idx == len(area_locations) and is_running:
        console_logger.info(f"\n🎉 区域{target_area}所有任务执行完毕！")
    else:
        console_logger.info(f"\n⚠️  任务执行被中断（已执行{idx-1}/{len(area_locations)}个任务）")
    
    if file_logger:
        file_logger.info(
            f"规则1执行结束：总执行{task_stats.total}个，成功{task_stats.success}个，"
            f"失败{task_stats.fail}个，中断状态={not is_running}，最终连续失败次数={task_stats.continuous_failures}"
        )

def run_rule2(
    grouped_areas: Dict[str, List[LocationInfo]],
    valid_areas: List[str],
    config: TaskConfig
):
    """规则2：多个区域随机执行（包含连续失败检查）"""
    global is_running
    is_running = True
    session = init_http_session(config)

    # 初始化任务队列（复制列表避免修改原数据）
    area_tasks = {area: grouped_areas[area].copy() for area in valid_areas}
    remaining_areas = [area for area in valid_areas if area_tasks[area]]

    console_logger.info(f"\n🚀 开始执行任务（规则2：随机调用）")
    if file_logger:
        file_logger.info(
            f"开始执行规则2：有效区域={sorted(valid_areas)}（来自[business]节areas），区域任务数={ {k:len(v) for k,v in area_tasks.items()} }，"
            f"store模式={'固定：' + config.fixed_store if config.fixed_store else '随机'}，"
            f"请求配置：超时={config.request_timeout}s，重试={config.retry_count}次，"
            f"成功条件：success=true 或 error_id={TARGET_ERROR_ID}"
        )

    try:
        while remaining_areas and is_running:
            # 检查连续失败次数，达到阈值则停止
            if check_continuous_failures():
                return

            # 随机选择区域
            current_area = random.choice(remaining_areas)
            current_loc = area_tasks[current_area].pop(0)
            
            # 选择store
            selected_store = config.fixed_store if config.fixed_store else random.choice(config.locations)
            
            # 任务序号
            task_seq = task_stats.total + 1
            
            # 控制台进度显示
            console_logger.info(f"\n[{task_seq}/{task_stats.total_task_count}] 区域：{current_loc.full_area} | ID：{current_loc.location_id}")
            
            # 发送任务
            success = send_task(session, current_loc.location_id, selected_store, config)

            # 更新统计
            if success:
                task_stats.success += 1
                console_logger.info(f"  ✅ 执行成功 | 目标store：{selected_store}")
            else:
                task_stats.fail += 1
            task_stats.total += 1

            # 更新剩余区域列表
            remaining_areas = [area for area in valid_areas if area_tasks[area]]
            
            # 任务间隔
            if remaining_areas:
                time.sleep(0.5)

    except Exception as e:
        console_logger.error(f"\n❌ 任务执行异常：{str(e)}")
        if file_logger:
            file_logger.error(f"规则2执行异常", exc_info=True)
    finally:
        is_running = False

    # 执行结果
    if not remaining_areas and is_running:
        console_logger.info(f"\n🎉 所有任务执行完毕！")
    else:
        remaining_task_count = sum(len(tasks) for tasks in area_tasks.values())
        console_logger.info(f"\n⚠️  任务执行被中断（已执行{task_stats.total}个，剩余{remaining_task_count}个）")
    
    if file_logger:
        file_logger.info(
            f"规则2执行结束：总执行{task_stats.total}个，成功{task_stats.success}个，"
            f"失败{task_stats.fail}个，剩余任务数={sum(len(tasks) for tasks in area_tasks.values())}，"
            f"最终连续失败次数={task_stats.continuous_failures}"
        )

# ==============================
# 报告和清理函数
# ==============================
def print_final_report():
    """输出最终报告（包含连续失败次数）"""
    if not task_stats.start_time:
        return

    # 计算执行时长
    end_time = task_stats.end_time or datetime.now()
    duration = (end_time - task_stats.start_time).total_seconds()
    duration_str = str(timedelta(seconds=duration)).split('.')[0]  # 格式化时长

    # 计算成功率
    success_rate = (task_stats.success / task_stats.total * 100) if task_stats.total > 0 else 0.0

    # 控制台简洁报告
    console_logger.info(f"\n" + "="*60)
    console_logger.info(f"🎯 任务执行报告")
    console_logger.info(f"="*60)
    console_logger.info(f"启动时间：{task_stats.start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    console_logger.info(f"结束时间：{end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    console_logger.info(f"执行时长：{duration_str}")
    console_logger.info(f"接口主机：{global_config.host}:{global_config.port}")
    console_logger.info(f"Excel工作表：{global_config.sheet_name or '最新工作表（最后一个）'}")
    console_logger.info(f"执行规则：{global_config.rule}")
    console_logger.info(f"执行区域（来自[business]节）：{sorted(task_stats.valid_areas)}")
    console_logger.info(f"Store模式：{'固定：' + global_config.fixed_store if global_config.fixed_store else '随机'}")
    console_logger.info(f"请求配置：超时={global_config.request_timeout}s，重试={global_config.retry_count}次")
    console_logger.info(f"成功条件：success=true 或 error_id={TARGET_ERROR_ID}")
    console_logger.info(f"="*60)
    console_logger.info(f"总待执行任务：{task_stats.total_task_count} 个")
    console_logger.info(f"已执行任务：{task_stats.total} 个")
    console_logger.info(f"成功任务：{task_stats.success} 个（{success_rate:.1f}%）")
    console_logger.info(f"失败任务：{task_stats.fail} 个（{100-success_rate:.1f}%）")
    console_logger.info(f"跳过任务：{task_stats.skipped} 个")
    console_logger.info(f"最大连续失败次数：{task_stats.continuous_failures}/{MAX_CONTINUOUS_FAILURES}")
    console_logger.info(f"="*60)

    # 文件详细报告
    if file_logger:
        file_logger.info(f"\n" + "="*80)
        file_logger.info(f"任务执行最终报告")
        file_logger.info(f"="*80)
        file_logger.info(f"启动时间：{task_stats.start_time.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]}")
        file_logger.info(f"结束时间：{end_time.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]}")
        file_logger.info(f"执行时长：{duration:.3f} 秒")
        file_logger.info(f"接口主机：{global_config.host}:{global_config.port}")
        file_logger.info(f"Excel文件：{global_config.xlsx_path}")
        file_logger.info(f"Excel工作表：{global_config.sheet_name or '最新工作表（最后一个）'}")
        file_logger.info(f"执行规则：{global_config.rule}")
        file_logger.info(f"执行区域（来自[business]节areas）：{sorted(task_stats.valid_areas)}")
        file_logger.info(f"Store模式：{'固定：' + global_config.fixed_store if global_config.fixed_store else '随机'}")
        file_logger.info(f"场景ID：{global_config.scene_id}")
        file_logger.info(f"请求配置：超时={global_config.request_timeout}s，重试={global_config.retry_count}次，延迟={global_config.retry_delay}s")
        file_logger.info(f"成功条件：success=true 或 error_id={TARGET_ERROR_ID}")
        file_logger.info(f"最大连续失败阈值：{MAX_CONTINUOUS_FAILURES}次")
        file_logger.info(f"默认错误信息：{DEFAULT_ERROR_INFO}")
        file_logger.info(f"="*80)
        file_logger.info(f"总待执行任务数：{task_stats.total_task_count} 个")
        file_logger.info(f"已执行任务数：{task_stats.total} 个")
        file_logger.info(f"成功任务数：{task_stats.success} 个（{success_rate:.2f}%）")
        file_logger.info(f"失败任务数：{task_stats.fail} 个（{100-success_rate:.2f}%）")
        file_logger.info(f"跳过任务数：{task_stats.skipped} 个")
        file_logger.info(f"最终连续失败次数：{task_stats.continuous_failures}")
        file_logger.info(f"="*80)

def cleanup_resources():
    """清理资源（关闭Excel工作簿、HTTP会话）"""
    global resource_handles
    if file_logger:
        file_logger.info("开始清理资源...")

    # 关闭Excel工作簿
    if resource_handles.get("excel_workbook"):
        try:
            resource_handles["excel_workbook"].close()
            if file_logger:
                file_logger.info("Excel工作簿已关闭")
        except Exception as e:
            if file_logger:
                file_logger.error(f"关闭Excel工作簿失败：{str(e)}")

    # 关闭HTTP会话
    if resource_handles.get("http_session"):
        try:
            resource_handles["http_session"].close()
            if file_logger:
                file_logger.info("HTTP会话已关闭")
        except Exception as e:
            if file_logger:
                file_logger.error(f"关闭HTTP会话失败：{str(e)}")

    # 重置资源句柄
    resource_handles = {"excel_workbook": None, "http_session": None}
    if file_logger:
        file_logger.info("资源清理完成")

def signal_handler(signum, frame):
    """信号处理函数（处理Ctrl+C中断）"""
    global is_running
    if not is_running:
        console_logger.info("\n⚠️  程序正在退出...")
        cleanup_resources()
        sys.exit(1)
    
    console_logger.info("\n⚠️  收到中断信号，正在停止任务...（再次按下Ctrl+C将强制退出）")
    is_running = False

# ==============================
# 主函数
# ==============================
def main():
    global global_config, task_stats

    # 注册信号处理（Ctrl+C）
    signal.signal(signal.SIGINT, signal_handler)

    # 解析命令行参数
    parser = argparse.ArgumentParser(description="区域取货到提升机任务调用工具")
    parser.add_argument("--config", type=str, default="./config.ini", help="配置文件路径（默认：./config.ini）")
    parser.add_argument("--rule", type=int, choices=VALID_RULES, help=f"执行规则（1=单个区域顺序；2=多个区域随机），覆盖配置文件")
    parser.add_argument("--areas", type=str, nargs="+", help="目标区域列表（空格分隔），覆盖配置文件")
    parser.add_argument("--fixed-store", type=str, help="固定目标store，覆盖配置文件")
    parser.add_argument("--debug", action="store_true", help="开启调试模式（覆盖配置文件log.debug）")
    args = parser.parse_args()

    try:
        # 1. 先加载日志配置（独立于完整配置）
        log_file, debug = load_log_config_from_ini(args.config)
        if not log_file:
            sys.exit(1)
        # 应用命令行debug参数
        if args.debug:
            debug = True
        init_loggers(log_file, debug)

        # 2. 加载完整配置
        global_config = load_full_config(args.config, args)
        if not global_config:
            cleanup_resources()
            sys.exit(1)

        # 3. 加载Excel数据
        location_list = load_xlsx_data(global_config.xlsx_path, global_config.sheet_name)
        if not location_list:
            console_logger.error("❌ Excel无有效数据，程序退出")
            cleanup_resources()
            sys.exit(1)

        # 4. 区域分组与验证
        grouped_areas, valid_areas = group_and_validate_areas(location_list, global_config.areas)
        if not grouped_areas or not valid_areas:
            console_logger.error("❌ 区域验证失败，程序退出")
            cleanup_resources()
            sys.exit(1)

        # 5. 初始化任务统计
        task_stats.start_time = datetime.now()

        # 6. 执行对应规则的任务
        if global_config.rule == 1:
            run_rule1(grouped_areas, valid_areas, global_config)
        else:
            run_rule2(grouped_areas, valid_areas, global_config)

        # 7. 任务执行完成，输出最终报告
        task_stats.end_time = datetime.now()
        print_final_report()

        # 8. 清理资源
        cleanup_resources()

        # 9. 退出状态码
        if task_stats.continuous_failures >= MAX_CONTINUOUS_FAILURES:
            sys.exit(3)  # 连续失败退出
        elif task_stats.fail > 0:
            sys.exit(2)  # 有失败任务但未达连续阈值
        else:
            sys.exit(0)  # 全部成功

    except Exception as e:
        console_logger.error(f"\n❌ 程序执行异常：{str(e)}")
        if file_logger:
            file_logger.error("程序执行异常", exc_info=True)
        cleanup_resources()
        sys.exit(1)

if __name__ == "__main__":
    main()